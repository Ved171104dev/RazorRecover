from __future__ import annotations
import csv,hashlib,io,json,os,re,secrets,zipfile
from threading import Lock
from datetime import datetime,timedelta
from typing import Any
from sqlalchemy import func,select
from sqlalchemy.orm import Session
from app.db import *
from app.providers.razorpay_adapter import ProviderError,RazorpayAdapter
from app.services.crypto import decrypt_secret,encrypt_secret
from app.services.recovery import calculate_strategies,evaluate_policy

RECOVERY_MODEL:Any=None
RECOVERY_MODEL_LOCK=Lock()

def recovery_model():
    global RECOVERY_MODEL
    if RECOVERY_MODEL is None:
        with RECOVERY_MODEL_LOCK:
            if RECOVERY_MODEL is None:
                from app.ml.inference import RecoveryModel
                RECOVERY_MODEL=RecoveryModel()
    return RECOVERY_MODEL

def mask_key(key_id:str)->str:return key_id[:9]+"********"

def ensure_webhook_token(connection:RazorpayConnection)->str:
    if not connection.webhook_token:connection.webhook_token=secrets.token_urlsafe(24)
    return connection.webhook_token

def provider_for_merchant(db:Session,merchant_id:str)->RazorpayAdapter:
    connection=db.scalar(select(RazorpayConnection).where(RazorpayConnection.merchant_id==merchant_id))
    if connection and connection.connection_status=="connected" and connection.key_id_encrypted and connection.key_secret_encrypted:
        return RazorpayAdapter(decrypt_secret(connection.key_id_encrypted),decrypt_secret(connection.key_secret_encrypted))
    adapter=RazorpayAdapter(os.getenv("RAZORPAY_KEY_ID"),os.getenv("RAZORPAY_KEY_SECRET"))
    if not adapter.configured:raise ProviderError("Connect a verified Razorpay Test Mode account before executing recovery")
    return adapter

def connect_razorpay(db:Session,merchant_id:str,key_id:str,key_secret:str,webhook_secret:str)->RazorpayConnection:
    if not key_id.startswith("rzp_test_"):raise ValueError("Only Razorpay Test Mode keys beginning with rzp_test_ are accepted")
    if len(key_secret)<8 or len(webhook_secret)<8:raise ValueError("Key secret and webhook secret must each contain at least 8 characters")
    adapter=RazorpayAdapter(key_id,key_secret);adapter.verify_connection()
    connection=db.scalar(select(RazorpayConnection).where(RazorpayConnection.merchant_id==merchant_id))
    if not connection:connection=RazorpayConnection(merchant_id=merchant_id);db.add(connection)
    connection.key_id_masked=mask_key(key_id);connection.key_id_encrypted=encrypt_secret(key_id);connection.key_secret_encrypted=encrypt_secret(key_secret);connection.webhook_secret_encrypted=encrypt_secret(webhook_secret);connection.mode="test";connection.connection_status="connected";connection.webhook_status="pending";connection.last_verified_at=utcnow();connection.sync_error=None;ensure_webhook_token(connection)
    db.add(AuditLog(merchant_id=merchant_id,actor_type="merchant",event_type="razorpay_connected",detail={"message":"Razorpay Test Mode credentials verified","key_id":connection.key_id_masked},amount_paise=0));db.commit();return connection

def disconnect_razorpay(db:Session,merchant_id:str)->None:
    connection=db.scalar(select(RazorpayConnection).where(RazorpayConnection.merchant_id==merchant_id))
    if not connection:return
    connection.key_id_encrypted=None;connection.key_secret_encrypted=None;connection.webhook_secret_encrypted=None;connection.key_id_masked=None;connection.connection_status="not_connected";connection.webhook_status="not_verified";connection.mode="not_connected";connection.sync_status="never";connection.sync_error=None
    db.add(AuditLog(merchant_id=merchant_id,actor_type="merchant",event_type="razorpay_disconnected",detail={"message":"Stored Razorpay credentials removed"},amount_paise=0));db.commit()

def _dt(value:Any)->datetime:
    try:return datetime.utcfromtimestamp(int(value))
    except Exception:return utcnow()

def _customer(db:Session,merchant_id:str,ref:str,email:str|None,name:str|None,phone:str|None)->Customer:
    safe_ref=(ref or hashlib.sha256((email or name or "unknown").encode()).hexdigest()[:24])[:120]
    row=db.scalar(select(Customer).where(Customer.merchant_id==merchant_id,Customer.external_ref==safe_ref))
    if not row:
        safe_email=(email or f"unknown+{safe_ref[:18]}@import.local")[:320]
        row=Customer(merchant_id=merchant_id,external_ref=safe_ref,name=(name or "Razorpay Customer")[:160],email=safe_email,phone=(phone or None),historical_success_rate=.65,preferred_method="card",previous_failures=0);db.add(row);db.flush()
    else:
        if email:row.email=email[:320]
        if name:row.name=name[:160]
        if phone is not None:row.phone=phone[:32] or None
    return row

def _order(db:Session,merchant_id:str,external_ref:str,customer:Customer,amount:int,currency:str,status:str,source:str,created_at:datetime)->Order:
    row=db.scalar(select(Order).where(Order.merchant_id==merchant_id,Order.external_ref==external_ref))
    if not row:
        row=Order(merchant_id=merchant_id,customer_id=customer.id,external_ref=external_ref[:120],amount_paise=amount,currency=currency[:3],status=status,data_source=source,created_at=created_at);db.add(row);db.flush()
    else:row.amount_paise=amount;row.currency=currency[:3];row.status=status;row.customer_id=customer.id
    return row

def _analyse_failure(db:Session,merchant_id:str,payment:Payment,customer:Customer)->None:
    if payment.status!="failed" or db.scalar(select(RiskEvent).where(RiskEvent.merchant_id==merchant_id,RiskEvent.payment_id==payment.id)):return
    code=(payment.failure_code or "PAYMENT_FAILURE").upper();retry_count=max(0,(db.scalar(select(func.count()).select_from(PaymentAttempt).where(PaymentAttempt.payment_id==payment.id)) or 1)-1)
    prediction=recovery_model().predict({"amount_paise":payment.amount_paise,"method":payment.method or "unknown","failure_code":code,"retry_count":retry_count,"historical_success":customer.historical_success_rate,"preferred_method":customer.preferred_method,"device":"unknown"})
    probability=prediction["recovery_probability"];confidence=prediction["confidence"];reason_codes=list(dict.fromkeys([code,*prediction["reason_codes"]]));mandate_window=payment.payment_type=="recurring" and payment.created_at>=utcnow()-timedelta(days=7)
    risk=RiskEvent(merchant_id=merchant_id,payment_id=payment.id,risk_score=prediction["risk_score"],recovery_probability=probability,affected_revenue_paise=payment.amount_paise,confidence=confidence,root_cause=code,reason_codes=reason_codes,evidence=[{"signal":"provider_failure","value":code,"detail":"Failure information supplied by the merchant data source."},{"signal":"payment_type","value":payment.payment_type,"detail":"Payment type determines whether provider-managed retry or customer consent is required."},{"signal":"local_model","value":prediction["model_version"],"detail":f"Local inference completed in {prediction['inference_latency_ms']} ms; no LLM participated in the decision."}]);db.add(risk);db.flush()
    policy=db.scalar(select(MerchantPolicy).where(MerchantPolicy.merchant_id==merchant_id));strategies=calculate_strategies(payment.amount_paise,probability,code,customer.preferred_method,retry_count,confidence,payment_type=payment.payment_type,payment_method=payment.method,mandate_grace_window_active=mandate_window);chosen=strategies[0];pr=evaluate_policy(amount_paise=payment.amount_paise,retry_count=retry_count,confidence=confidence,action=chosen["action"],allowed_actions=policy.allowed_actions,automatic_threshold_paise=policy.automatic_threshold_paise,approval_threshold_paise=policy.approval_threshold_paise,blocked_threshold_paise=policy.blocked_threshold_paise,max_retries=policy.max_retries,minimum_confidence=policy.minimum_confidence,payment_type=payment.payment_type,payment_method=payment.method,failure_code=code,mandate_grace_window_active=mandate_window)
    db.add(AgentDecision(merchant_id=merchant_id,risk_event_id=risk.id,selected_action=chosen["action"],candidates=strategies,expected_recovery_paise=chosen["expected_recovery_paise"],predicted_probability=chosen["probability"],confidence=confidence,policy_result={**pr,"decision_engine":prediction["decision_engine"],"inference_latency_ms":prediction["inference_latency_ms"]},policy_status="approval_required" if pr["approval_required"] else ("approved" if pr["allowed"] else "blocked"),explanation=chosen["reason"],model_version=prediction["model_version"]))

def _upsert_payment(db:Session,merchant_id:str,item:dict,order:Order,customer:Customer,source:str)->Payment:
    pid=str(item.get("id") or item.get("external_id") or "")[:120]
    row=db.scalar(select(Payment).where(Payment.merchant_id==merchant_id,Payment.external_ref==pid))
    amount=int(item.get("amount") or item.get("amount_paise") or order.amount_paise);status=str(item.get("status") or "failed");method=item.get("method");error=item.get("error_code") or item.get("failure_code");payment_type="recurring" if item.get("recurring") or item.get("subscription_id") or item.get("payment_type")=="recurring" else "one_time"
    if not row:
        row=Payment(merchant_id=merchant_id,order_id=order.id,external_ref=pid,amount_paise=amount,currency=str(item.get("currency") or "INR")[:3],method=method,payment_type=payment_type,failure_code=error,failure_description=item.get("error_description"),bank=item.get("bank"),status=status,data_source=source,created_at=_dt(item.get("created_at")));db.add(row);db.flush()
    else:row.order_id=order.id;row.amount_paise=amount;row.currency=str(item.get("currency") or "INR")[:3];row.status=status;row.method=method;row.payment_type=payment_type;row.failure_code=error;row.failure_description=item.get("error_description");row.data_source=source
    attempt=db.scalar(select(PaymentAttempt).where(PaymentAttempt.merchant_id==merchant_id,PaymentAttempt.payment_id==row.id,PaymentAttempt.attempt_number==1))
    if not attempt:db.add(PaymentAttempt(merchant_id=merchant_id,payment_id=row.id,attempt_number=1,method=method or "unknown",status=status,failure_code=error,device="unknown",checkout_duration_seconds=0,created_at=row.created_at))
    else:attempt.method=method or "unknown";attempt.status=status;attempt.failure_code=error
    _analyse_failure(db,merchant_id,row,customer);return row

def sync_razorpay(db:Session,merchant_id:str,days:int=30,max_records:int=1000)->DataIngestionRun:
    connection=db.scalar(select(RazorpayConnection).where(RazorpayConnection.merchant_id==merchant_id))
    if not connection or connection.connection_status!="connected":raise ValueError("Connect and verify Razorpay Test Mode before synchronization")
    run=DataIngestionRun(merchant_id=merchant_id,source="razorpay_test",idempotency_key=f"sync:{utcnow().isoformat()}",status="running",counts={});db.add(run);connection.sync_status="running";db.commit()
    try:
        adapter=provider_for_merchant(db,merchant_id);since=int((utcnow()-timedelta(days=days)).timestamp());orders=adapter.list_orders(since,max_records);payments=adapter.list_payments(since,max_records);order_map:dict[str,Order]={}
        for item in orders:
            oid=str(item.get("id"));notes=item.get("notes") if isinstance(item.get("notes"),dict) else {};customer=_customer(db,merchant_id,str(notes.get("customer_id") or f"order:{oid}"),notes.get("email"),notes.get("name"),notes.get("contact"));order=_order(db,merchant_id,oid,customer,int(item.get("amount") or 0),str(item.get("currency") or "INR"),str(item.get("status") or "created"),"razorpay_test",_dt(item.get("created_at")));order_map[oid]=order
            rp=db.scalar(select(RazorpayOrder).where(RazorpayOrder.merchant_id==merchant_id,RazorpayOrder.razorpay_order_id==oid))
            if not rp:db.add(RazorpayOrder(merchant_id=merchant_id,internal_order_id=order.id,razorpay_order_id=oid,amount_paise=order.amount_paise,currency=order.currency,status=order.status,raw_data=item))
            else:rp.status=order.status;rp.raw_data=item
        for item in payments:
            pid=str(item.get("id"));oid=str(item.get("order_id") or f"standalone:{pid}");customer=_customer(db,merchant_id,str(item.get("customer_id") or item.get("email") or f"payment:{pid}"),item.get("email"),None,item.get("contact"));order=order_map.get(oid) or _order(db,merchant_id,oid,customer,int(item.get("amount") or 0),str(item.get("currency") or "INR"),"paid" if item.get("status")=="captured" else "payment_failed", "razorpay_test",_dt(item.get("created_at")));payment=_upsert_payment(db,merchant_id,item,order,customer,"razorpay_test")
            rp=db.scalar(select(RazorpayPayment).where(RazorpayPayment.merchant_id==merchant_id,RazorpayPayment.razorpay_payment_id==pid))
            if not rp:db.add(RazorpayPayment(merchant_id=merchant_id,internal_payment_id=payment.id,razorpay_payment_id=pid,razorpay_order_id=item.get("order_id"),amount_paise=payment.amount_paise,currency=payment.currency,status=payment.status,method=payment.method,failure_code=payment.failure_code,raw_data=item))
            else:rp.internal_payment_id=payment.id;rp.status=payment.status;rp.raw_data=item
        counts={"orders":len(orders),"payments":len(payments),"failed_payments":sum(1 for x in payments if x.get("status")=="failed")};run.status="completed";run.counts=counts;run.completed_at=utcnow();connection.last_sync_at=utcnow();connection.sync_status="completed";connection.sync_error=None;connection.imported_orders=len(orders);connection.imported_payments=len(payments);db.add(AuditLog(merchant_id=merchant_id,event_type="razorpay_sync_completed",detail={"message":"Razorpay Test Mode synchronization completed",**counts},amount_paise=0));db.commit();return run
    except Exception as exc:
        db.rollback();run=db.get(DataIngestionRun,run.id);connection=db.scalar(select(RazorpayConnection).where(RazorpayConnection.merchant_id==merchant_id));run.status="failed";run.error=str(exc)[:500];run.completed_at=utcnow();connection.sync_status="failed";connection.sync_error=str(exc)[:500];db.commit();raise

IMPORT_COLUMNS={"external_id","order_id","customer_email","customer_name","amount_paise","status","method","failure_code"}
SUPPORTED_IMPORT_EXTENSIONS={".csv",".tsv",".xlsx",".xls",".json",".pdf"}
HEADER_ALIASES={"payment_id":"external_id","paymentid":"external_id","order":"order_id","email":"customer_email","customer":"customer_name","amount":"amount_paise","error_code":"failure_code","failure_reason":"failure_code","phone":"customer_phone"}
MAX_IMPORT_BYTES=10_000_000
MAX_IMPORT_ROWS=5000
FILE_IMPORT_SOURCES={"merchant_file","merchant_csv"}

def _header(value:Any)->str:
    normal=re.sub(r"[^a-z0-9]+","_",str(value or "").strip().lower()).strip("_")
    return HEADER_ALIASES.get(normal,normal)

def _cell(value:Any)->str:
    if value is None:return ""
    if isinstance(value,float) and value.is_integer():return str(int(value))
    return str(value).strip()

def _normalise_dicts(records:list[dict[Any,Any]])->list[dict[str,str]]:
    output=[]
    for raw in records:
        item={_header(k):_cell(v) for k,v in raw.items() if _header(k)}
        if any(item.values()):output.append(item)
        if len(output)>MAX_IMPORT_ROWS:raise ValueError(f"File contains more than {MAX_IMPORT_ROWS:,} payment rows")
    missing=IMPORT_COLUMNS-set(output[0] if output else {})
    if missing:raise ValueError("Payment table is missing columns: "+", ".join(sorted(missing)))
    return output

def _matrix_records(matrix:Any)->list[dict[str,str]]:
    headers=None;records=[]
    for index,row in enumerate(matrix):
        values=list(row or [])
        candidate=[_header(x) for x in values]
        if headers is None:
            if IMPORT_COLUMNS.issubset(set(candidate)):headers=candidate
            elif index>=24:break
            continue
        if IMPORT_COLUMNS.issubset(set(candidate)):continue
        if not any(_cell(x) for x in values):continue
        records.append({headers[i]:_cell(values[i]) if i<len(values) else "" for i in range(len(headers)) if headers[i]})
        if len(records)>MAX_IMPORT_ROWS:raise ValueError(f"File contains more than {MAX_IMPORT_ROWS:,} payment rows")
    return _normalise_dicts(records)

def _decode_text(raw:bytes,label:str)->str:
    try:return raw.decode("utf-8-sig")
    except UnicodeDecodeError as exc:raise ValueError(f"{label} must use UTF-8 encoding") from exc

def _parse_delimited(raw:bytes,delimiter:str,label:str)->list[dict[str,str]]:
    reader=csv.DictReader(io.StringIO(_decode_text(raw,label)),delimiter=delimiter)
    return _normalise_dicts(list(reader))

def _check_xlsx_archive(raw:bytes)->None:
    if not raw.startswith(b"PK"):raise ValueError("XLSX file signature is invalid")
    try:
        with zipfile.ZipFile(io.BytesIO(raw)) as archive:
            if len(archive.infolist())>2000 or sum(x.file_size for x in archive.infolist())>50_000_000:raise ValueError("XLSX archive is too large after decompression")
    except zipfile.BadZipFile as exc:raise ValueError("XLSX file is damaged or invalid") from exc

def _parse_xlsx(raw:bytes)->list[dict[str,str]]:
    _check_xlsx_archive(raw)
    try:
        from openpyxl import load_workbook
        workbook=load_workbook(io.BytesIO(raw),read_only=True,data_only=True)
        for sheet in workbook.worksheets:
            try:return _matrix_records(sheet.iter_rows(values_only=True))
            except ValueError as exc:
                if "missing columns" not in str(exc):raise
        raise ValueError("No worksheet contains the required payment columns")
    except ValueError:raise
    except Exception as exc:raise ValueError("XLSX file could not be read") from exc

def _parse_xls(raw:bytes)->list[dict[str,str]]:
    if not raw.startswith(bytes.fromhex("D0CF11E0")):raise ValueError("XLS file signature is invalid")
    try:
        import xlrd
        workbook=xlrd.open_workbook(file_contents=raw,on_demand=True)
        for sheet in workbook.sheets():
            try:return _matrix_records(([cell.value for cell in row] for row in sheet.get_rows()))
            except ValueError as exc:
                if "missing columns" not in str(exc):raise
        raise ValueError("No worksheet contains the required payment columns")
    except ValueError:raise
    except Exception as exc:raise ValueError("XLS file could not be read") from exc

def _parse_json(raw:bytes)->list[dict[str,str]]:
    try:data=json.loads(_decode_text(raw,"JSON"))
    except json.JSONDecodeError as exc:raise ValueError("JSON file is malformed") from exc
    records=data.get("payments") if isinstance(data,dict) else data
    if not isinstance(records,list) or not all(isinstance(x,dict) for x in records):raise ValueError("JSON must be an array of payment objects or an object containing a payments array")
    return _normalise_dicts(records)

def _parse_pdf(raw:bytes)->list[dict[str,str]]:
    if not raw.startswith(b"%PDF"):raise ValueError("PDF file signature is invalid")
    try:
        import pdfplumber
        records=[]
        with pdfplumber.open(io.BytesIO(raw)) as document:
            if len(document.pages)>50:raise ValueError("PDF contains more than 50 pages")
            for page in document.pages:
                for table in page.extract_tables() or []:
                    try:records.extend(_matrix_records(table))
                    except ValueError as exc:
                        if "missing columns" not in str(exc):raise
        if not records:raise ValueError("PDF must contain a machine-readable table with the required payment columns; scanned PDFs are not supported")
        if len(records)>MAX_IMPORT_ROWS:raise ValueError(f"File contains more than {MAX_IMPORT_ROWS:,} payment rows")
        return records
    except ValueError:raise
    except Exception as exc:raise ValueError("PDF could not be read as a payment table") from exc

def parse_payment_file(filename:str,raw:bytes)->tuple[list[dict[str,str]],str]:
    extension=os.path.splitext(filename.lower())[1]
    if extension not in SUPPORTED_IMPORT_EXTENSIONS:raise ValueError("Supported files: CSV, TSV, XLSX, XLS, JSON, and table-based PDF")
    if not raw:raise ValueError("Uploaded file is empty")
    if len(raw)>MAX_IMPORT_BYTES:raise ValueError("File exceeds the 10 MB limit")
    parsers={".csv":lambda:_parse_delimited(raw,",","CSV"),".tsv":lambda:_parse_delimited(raw,"\t","TSV"),".xlsx":lambda:_parse_xlsx(raw),".xls":lambda:_parse_xls(raw),".json":lambda:_parse_json(raw),".pdf":lambda:_parse_pdf(raw)}
    rows=parsers[extension]()
    if not rows:raise ValueError("File does not contain payment rows")
    return rows,extension.lstrip(".")

def normalise_import_record(item:dict[str,Any],index:int=2)->dict[str,Any]:
    amount_text=str(item.get("amount_paise") or "").replace(",","").strip()
    try:amount=int(amount_text)
    except Exception as exc:raise ValueError(f"Row {index}: amount_paise must be an integer") from exc
    if amount<=0:raise ValueError(f"Row {index}: amount_paise must be positive")
    external=str(item.get("external_id") or "").strip()
    if not external:raise ValueError(f"Row {index}: external_id is required")
    raw_status=str(item.get("status") or "failed").strip().lower()
    status={"paid":"captured","success":"captured","successful":"captured","declined":"failed"}.get(raw_status,raw_status)
    if status not in {"captured","authorized","failed"}:raise ValueError(f"Row {index}: status must be captured, authorized, or failed")
    email=str(item.get("customer_email") or "").strip()
    name=str(item.get("customer_name") or "").strip()
    order_id=str(item.get("order_id") or f"file-order:{external}").strip()
    method=str(item.get("method") or "").strip().lower()
    failure_code=str(item.get("failure_code") or "").strip().upper()
    currency=str(item.get("currency") or "INR").strip().upper()
    phone=str(item.get("customer_phone") or "").strip();payment_type=str(item.get("payment_type") or "one_time").strip().lower()
    if payment_type not in {"one_time","recurring"}:raise ValueError(f"Row {index}: payment_type must be one_time or recurring")
    return {"external_id":external[:120],"order_id":order_id[:120],"customer_email":email[:320],"customer_name":name[:160],"amount_paise":amount,"status":status,"method":method[:32],"failure_code":failure_code[:80],"currency":currency[:3],"customer_phone":phone[:32],"payment_type":payment_type}

def _assert_payment_editable(db:Session,merchant_id:str,payment:Payment)->None:
    actions=db.scalars(select(RecoveryAction).where(RecoveryAction.merchant_id==merchant_id,RecoveryAction.payment_id==payment.id)).all()
    for action in actions:
        attributed=db.scalar(select(RecoveryAttribution).where(RecoveryAttribution.merchant_id==merchant_id,RecoveryAttribution.recovery_action_id==action.id))
        provider_link=db.scalar(select(RazorpayPaymentLink).where(RazorpayPaymentLink.merchant_id==merchant_id,RazorpayPaymentLink.recovery_action_id==action.id))
        if attributed or provider_link or action.provider_reference or action.provider_url or action.razorpay_payment_id or action.executed_at or action.verified_at or action.verification_status!="not_started" or action.actual_recovered_paise:
            raise ValueError("This payment has an executed or verified recovery and cannot be edited or removed. Preserve it for financial audit integrity.")

def _discard_unexecuted_actions(db:Session,merchant_id:str,payment:Payment)->None:
    _assert_payment_editable(db,merchant_id,payment)
    actions=db.scalars(select(RecoveryAction).where(RecoveryAction.merchant_id==merchant_id,RecoveryAction.payment_id==payment.id)).all()
    for action in actions:
        for approval in db.scalars(select(Approval).where(Approval.merchant_id==merchant_id,Approval.recovery_action_id==action.id)).all():db.delete(approval)
        for result in db.scalars(select(ExperimentResult).where(ExperimentResult.merchant_id==merchant_id,ExperimentResult.recovery_action_id==action.id)).all():db.delete(result)
        for audit in db.scalars(select(AuditLog).where(AuditLog.merchant_id==merchant_id,AuditLog.action_id==action.id)).all():audit.action_id=None
        for event in db.scalars(select(AgentEvent).where(AgentEvent.merchant_id==merchant_id,AgentEvent.action_id==action.id)).all():event.action_id=None
        db.delete(action)
    db.flush()

def _clear_failure_analysis(db:Session,merchant_id:str,payment:Payment)->None:
    _discard_unexecuted_actions(db,merchant_id,payment)
    risks=db.scalars(select(RiskEvent).where(RiskEvent.merchant_id==merchant_id,RiskEvent.payment_id==payment.id)).all()
    for risk in risks:
        decisions=db.scalars(select(AgentDecision).where(AgentDecision.merchant_id==merchant_id,AgentDecision.risk_event_id==risk.id)).all()
        for decision in decisions:
            for audit in db.scalars(select(AuditLog).where(AuditLog.merchant_id==merchant_id,AuditLog.decision_id==decision.id)).all():audit.decision_id=None
            db.delete(decision)
        db.flush();db.delete(risk)
    db.flush()

def _materialize_import_record(db:Session,merchant_id:str,item:dict[str,Any],refresh:bool=True)->Payment:
    existing=db.scalar(select(Payment).where(Payment.merchant_id==merchant_id,Payment.external_ref==item["external_id"]))
    if existing and refresh:_clear_failure_analysis(db,merchant_id,existing)
    customer=_customer(db,merchant_id,item["customer_email"] or f"file:{item['external_id']}",item["customer_email"] or None,item["customer_name"] or None,item["customer_phone"] or None)
    order=_order(db,merchant_id,item["order_id"],customer,item["amount_paise"],item["currency"],"paid" if item["status"] in {"captured","authorized"} else "payment_failed","merchant_import",utcnow())
    return _upsert_payment(db,merchant_id,{"id":item["external_id"],"amount":item["amount_paise"],"currency":item["currency"],"status":item["status"],"method":item["method"],"failure_code":item["failure_code"] or None,"payment_type":item.get("payment_type","one_time")},order,customer,"merchant_import")

def _active_file_runs(db:Session,merchant_id:str,exclude_id:str|None=None)->list[DataIngestionRun]:
    query=select(DataIngestionRun).where(DataIngestionRun.merchant_id==merchant_id,DataIngestionRun.source.in_(FILE_IMPORT_SOURCES),DataIngestionRun.removed_at.is_(None)).order_by(DataIngestionRun.started_at.desc())
    if exclude_id:query=query.where(DataIngestionRun.id!=exclude_id)
    return list(db.scalars(query).all())

def _other_file_record(db:Session,merchant_id:str,external_id:str,exclude_id:str)->dict[str,Any]|None:
    for run in _active_file_runs(db,merchant_id,exclude_id):
        for record in run.records or []:
            if record.get("external_id")==external_id:return record
    return None

def _delete_import_payment(db:Session,merchant_id:str,external_id:str)->None:
    payment=db.scalar(select(Payment).where(Payment.merchant_id==merchant_id,Payment.external_ref==external_id,Payment.data_source=="merchant_import"))
    if not payment:return
    _clear_failure_analysis(db,merchant_id,payment)
    for attempt in db.scalars(select(PaymentAttempt).where(PaymentAttempt.merchant_id==merchant_id,PaymentAttempt.payment_id==payment.id)).all():db.delete(attempt)
    for provider_payment in db.scalars(select(RazorpayPayment).where(RazorpayPayment.merchant_id==merchant_id,RazorpayPayment.internal_payment_id==payment.id)).all():provider_payment.internal_payment_id=None
    db.flush();db.delete(payment)

def _update_run_counts(run:DataIngestionRun)->None:
    records=run.records or [];counts=dict(run.counts or {});counts["payments"]=len(records);counts["failed_payments"]=sum(1 for item in records if item.get("status")=="failed");counts["filename"]=run.filename;run.counts=counts

def backfill_legacy_import_records(db:Session,merchant_id:str)->None:
    runs=_active_file_runs(db,merchant_id)
    if not runs or any(run.records for run in runs):return
    payments=db.scalars(select(Payment).where(Payment.merchant_id==merchant_id,Payment.data_source=="merchant_import").order_by(Payment.created_at)).all()
    if not payments:return
    records=[]
    for payment in payments:
        order=db.get(Order,payment.order_id);customer=db.get(Customer,order.customer_id)
        records.append({"external_id":payment.external_ref,"order_id":order.external_ref,"customer_email":customer.email,"customer_name":customer.name,"amount_paise":payment.amount_paise,"status":payment.status,"method":payment.method or "","failure_code":payment.failure_code or "","currency":payment.currency,"customer_phone":customer.phone or "","payment_type":payment.payment_type})
    run=runs[0];run.records=records;run.filename=run.filename or (run.counts or {}).get("filename") or "legacy-merchant-import";_update_run_counts(run);db.commit()

def get_import_run(db:Session,merchant_id:str,run_id:str)->DataIngestionRun:
    run=db.scalar(select(DataIngestionRun).where(DataIngestionRun.id==run_id,DataIngestionRun.merchant_id==merchant_id,DataIngestionRun.source.in_(FILE_IMPORT_SOURCES),DataIngestionRun.removed_at.is_(None)))
    if not run:raise LookupError("Imported file not found")
    return run

def add_import_record(db:Session,merchant_id:str,run_id:str,item:dict[str,Any])->DataIngestionRun:
    run=get_import_run(db,merchant_id,run_id);record=normalise_import_record(item)
    if any(x.get("external_id")==record["external_id"] for x in run.records or []):raise ValueError("This file already contains that external_id")
    records=list(run.records or []);records.append(record);run.records=records;_materialize_import_record(db,merchant_id,record);_update_run_counts(run)
    db.add(AuditLog(merchant_id=merchant_id,actor_type="merchant",event_type="import_record_added",detail={"message":"Payment row added to imported file","filename":run.filename,"external_id":record["external_id"]},amount_paise=record["amount_paise"]));db.commit();return run

def update_import_record(db:Session,merchant_id:str,run_id:str,external_id:str,item:dict[str,Any])->DataIngestionRun:
    run=get_import_run(db,merchant_id,run_id);record=normalise_import_record(item);records=list(run.records or []);position=next((i for i,x in enumerate(records) if x.get("external_id")==external_id),None)
    if position is None:raise LookupError("Payment row not found")
    if record["external_id"]!=external_id and any(x.get("external_id")==record["external_id"] for x in records):raise ValueError("This file already contains the new external_id")
    old_payment=db.scalar(select(Payment).where(Payment.merchant_id==merchant_id,Payment.external_ref==external_id,Payment.data_source=="merchant_import"))
    if old_payment:_assert_payment_editable(db,merchant_id,old_payment)
    new_payment=db.scalar(select(Payment).where(Payment.merchant_id==merchant_id,Payment.external_ref==record["external_id"],Payment.data_source=="merchant_import"))
    if new_payment and new_payment.id!=(old_payment.id if old_payment else None):_assert_payment_editable(db,merchant_id,new_payment)
    records[position]=record;run.records=records
    if record["external_id"]!=external_id:
        other=_other_file_record(db,merchant_id,external_id,run.id)
        if other:_materialize_import_record(db,merchant_id,other)
        else:_delete_import_payment(db,merchant_id,external_id)
    _materialize_import_record(db,merchant_id,record);_update_run_counts(run)
    db.add(AuditLog(merchant_id=merchant_id,actor_type="merchant",event_type="import_record_updated",detail={"message":"Payment row updated in imported file","filename":run.filename,"external_id":record["external_id"],"previous_external_id":external_id},amount_paise=record["amount_paise"]));db.commit();return run

def remove_import_record(db:Session,merchant_id:str,run_id:str,external_id:str)->DataIngestionRun:
    run=get_import_run(db,merchant_id,run_id);records=list(run.records or []);record=next((x for x in records if x.get("external_id")==external_id),None)
    if not record:raise LookupError("Payment row not found")
    payment=db.scalar(select(Payment).where(Payment.merchant_id==merchant_id,Payment.external_ref==external_id,Payment.data_source=="merchant_import"))
    if payment:_assert_payment_editable(db,merchant_id,payment)
    run.records=[x for x in records if x.get("external_id")!=external_id]
    other=_other_file_record(db,merchant_id,external_id,run.id)
    if other:_materialize_import_record(db,merchant_id,other)
    else:_delete_import_payment(db,merchant_id,external_id)
    _update_run_counts(run);db.add(AuditLog(merchant_id=merchant_id,actor_type="merchant",event_type="import_record_removed",detail={"message":"Payment row removed from imported file","filename":run.filename,"external_id":external_id},amount_paise=record["amount_paise"]));db.commit();return run

def remove_import_run(db:Session,merchant_id:str,run_id:str)->None:
    run=get_import_run(db,merchant_id,run_id);records=list(run.records or [])
    for record in records:
        payment=db.scalar(select(Payment).where(Payment.merchant_id==merchant_id,Payment.external_ref==record["external_id"],Payment.data_source=="merchant_import"))
        if payment:_assert_payment_editable(db,merchant_id,payment)
    for record in records:
        other=_other_file_record(db,merchant_id,record["external_id"],run.id)
        if other:_materialize_import_record(db,merchant_id,other)
        else:_delete_import_payment(db,merchant_id,record["external_id"])
    previous_counts=dict(run.counts or {});run.records=[];run.removed_at=utcnow();run.status="removed";run.counts={**previous_counts,"payments":0,"failed_payments":0}
    db.add(AuditLog(merchant_id=merchant_id,actor_type="merchant",event_type="import_file_removed",detail={"message":"Imported file and its unreferenced payment rows were removed","filename":run.filename,"previous_counts":previous_counts},amount_paise=0));db.commit()

def import_payment_file(db:Session,merchant_id:str,filename:str,raw:bytes)->DataIngestionRun:
    digest=hashlib.sha256(raw).hexdigest();existing=db.scalar(select(DataIngestionRun).where(DataIngestionRun.merchant_id==merchant_id,DataIngestionRun.source=="merchant_file",DataIngestionRun.idempotency_key==digest))
    if existing and existing.removed_at is None:return existing
    records,file_type=parse_payment_file(filename,raw);records=[normalise_import_record(item,index) for index,item in enumerate(records,start=2)]
    safe_name=os.path.basename(filename)[:180]
    run=existing or DataIngestionRun(merchant_id=merchant_id,source="merchant_file",idempotency_key=digest,status="running",counts={});run.status="running";run.filename=safe_name;run.records=records;run.removed_at=None;run.error=None;db.add(run);db.flush();rows=0;failed=0
    try:
        for item in records:_materialize_import_record(db,merchant_id,item);rows+=1;failed+=int(item["status"]=="failed")
        run.status="completed";run.counts={"payments":rows,"failed_payments":failed,"format":file_type,"filename":safe_name};run.completed_at=utcnow();db.add(AuditLog(merchant_id=merchant_id,actor_type="merchant",event_type="file_import_completed",detail={"message":"Merchant payment file import completed","filename":safe_name,"format":file_type,"payments":rows,"failed_payments":failed,"sha256":digest},amount_paise=0));db.commit();return run
    except Exception as exc:db.rollback();raise ValueError(str(exc)) from exc

def import_csv(db:Session,merchant_id:str,raw:bytes)->DataIngestionRun:
    """Backward-compatible CSV import used by older API clients."""
    return import_payment_file(db,merchant_id,"payments.csv",raw)
